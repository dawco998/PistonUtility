import openpyxl
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def dump_xlsx(filepath, label):
    wb = openpyxl.load_workbook(filepath)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'='*80}")
        print(f"PLIK: {label} | ARKUSZ: {sheet_name}")
        print(f"Wymiary: {ws.dimensions} | Wiersze: {ws.max_row} | Kolumny: {ws.max_column}")
        print(f"{'='*80}")
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
            vals = []
            for cell in row:
                v = cell.value
                if v is None:
                    vals.append('')
                else:
                    vals.append(str(v))
            print(' | '.join(vals))

files = [
    (r'c:\Users\Dawi\Documents\MazdaEDC16C3\Duration_Map.xlsx', 'Duration_Map'),
    (r'c:\Users\Dawi\Documents\MazdaEDC16C3\NM_LIMITER_Map.xlsx', 'NM_LIMITER_Map'),
    (r'c:\Users\Dawi\Documents\MazdaEDC16C3\NM_TO_IQ_Map.xlsx', 'NM_TO_IQ_Map'),
    (r'c:\Users\Dawi\Documents\MazdaEDC16C3\Rail_Pressure_Map.xlsx', 'Rail_Pressure_Map'),
]

for fpath, label in files:
    dump_xlsx(fpath, label)
